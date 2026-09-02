"""PDF and Excel rendering shared by every report endpoint.

Both renderers work off the same simple shape every report already builds
for its JSON response: a list of (heading, content) sections, where content
is a flat dict (rendered as a two-column table), a list of dicts (rendered
as a table with the dict keys as headers), or a list of strings (rendered as
bullet points / one row). No report-specific formatting logic lives here --
each report endpoint decides its own sections, this module just lays them
out.
"""

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    ListFlowable, ListItem, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

Section = tuple[str, Any]


def _cell(value: Any) -> str:
    if value is None:
        return "n/a"
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, bool):
        return "yes" if value else "no"
    if isinstance(value, float):
        return f"{value:.4g}"
    return str(value)


def render_pdf(title: str, sections: list[Section]) -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    story: list[Any] = [Paragraph(title, styles["Title"]), Spacer(1, 0.5 * cm)]

    for heading, content in sections:
        story.append(Paragraph(heading, styles["Heading2"]))
        story.append(Spacer(1, 0.2 * cm))

        if isinstance(content, dict):
            rows = [[k, _cell(v)] for k, v in content.items()]
            if rows:
                story.append(_styled_table([["Field", "Value"]] + rows))
        elif isinstance(content, list) and content and isinstance(content[0], dict):
            headers = list(content[0].keys())
            rows = [[_cell(row.get(h)) for h in headers] for row in content]
            story.append(_styled_table([headers] + rows))
        elif isinstance(content, list):
            story.append(ListFlowable(
                [ListItem(Paragraph(_cell(item), styles["Normal"])) for item in content],
                bulletType="bullet",
            ))
        else:
            story.append(Paragraph(_cell(content), styles["Normal"]))
        story.append(Spacer(1, 0.4 * cm))

    doc.build(story)
    return buffer.getvalue()


def _styled_table(rows: list[list[str]]) -> Table:
    table = Table(rows, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#059669")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    return table


def render_xlsx(title: str, sections: list[Section]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True, color="FFFFFF")

    for heading, content in sections:
        sheet_name = heading[:31] or "Sheet"
        ws = wb.create_sheet(sheet_name)

        if isinstance(content, dict):
            ws.append(["Field", "Value"])
            for k, v in content.items():
                ws.append([k, _cell(v)])
        elif isinstance(content, list) and content and isinstance(content[0], dict):
            headers = list(content[0].keys())
            ws.append(headers)
            for row in content:
                ws.append([_cell(row.get(h)) for h in headers])
        elif isinstance(content, list):
            ws.append(["Item"])
            for item in content:
                ws.append([_cell(item)])
        else:
            ws.append(["Value"])
            ws.append([_cell(content)])

        for cell in ws[1]:
            cell.font = header_font
        for column_cells in ws.columns:
            width = max(len(_cell(c.value)) for c in column_cells) + 2
            ws.column_dimensions[column_cells[0].column_letter].width = min(width, 60)

    if not wb.sheetnames:
        wb.create_sheet("Report")

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

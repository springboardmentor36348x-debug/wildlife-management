"""
Reports & Export System (Milestone 4, spec section 13).

Generates PDF and Excel reports summarizing a monitoring site's species
observations, biodiversity assessment, population estimates, habitat
status, and conservation recommendations - the kind of document a forest
department officer would print or email, not just view in the dashboard.
"""
import io
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.auth.dependencies import get_current_user
from app.database import get_db
from app.models.user import User
from app.models.survey import MonitoringSite
from app.models.observation import SpeciesObservation, MediaAsset
from app.models.biodiversity import BiodiversityAssessment
from app.models.population import PopulationEstimate
from app.models.habitat import HabitatAssessment
from app.models.conservation import ConservationRecommendation

router = APIRouter(prefix="/api/v1/reports", tags=["Reports & Export System"])


def _gather_site_report_data(db: Session, monitoring_site_id: str):
    site = db.query(MonitoringSite).filter(MonitoringSite.id == monitoring_site_id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Monitoring site not found.")

    observations = (
        db.query(SpeciesObservation)
        .join(MediaAsset)
        .filter(MediaAsset.monitoring_site_id == monitoring_site_id)
        .order_by(SpeciesObservation.detected_at.desc())
        .all()
    )
    latest_biodiversity = (
        db.query(BiodiversityAssessment)
        .filter(BiodiversityAssessment.monitoring_site_id == monitoring_site_id)
        .order_by(BiodiversityAssessment.assessed_at.desc())
        .first()
    )
    latest_habitat = (
        db.query(HabitatAssessment)
        .filter(HabitatAssessment.monitoring_site_id == monitoring_site_id)
        .order_by(HabitatAssessment.assessed_at.desc())
        .first()
    )
    population_estimates = (
        db.query(PopulationEstimate)
        .filter(PopulationEstimate.monitoring_site_id == monitoring_site_id)
        .order_by(PopulationEstimate.assessed_at.desc())
        .all()
    )
    recommendations = (
        db.query(ConservationRecommendation)
        .filter(ConservationRecommendation.monitoring_site_id == monitoring_site_id)
        .order_by(ConservationRecommendation.generated_at.desc())
        .all()
    )

    return {
        "site": site,
        "observations": observations,
        "biodiversity": latest_biodiversity,
        "habitat": latest_habitat,
        "population_estimates": population_estimates,
        "recommendations": recommendations,
    }


@router.get("/{monitoring_site_id}/pdf")
def export_site_report_pdf(
    monitoring_site_id: str,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    data = _gather_site_report_data(db, monitoring_site_id)
    site = data["site"]

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleForest", parent=styles["Title"], textColor=colors.HexColor("#1C4D25"))
    heading_style = ParagraphStyle("HeadingForest", parent=styles["Heading2"], textColor=colors.HexColor("#1C4D25"))

    elements = [
        Paragraph("Wildlife Population Intelligence System", title_style),
        Paragraph(f"Site Report: {site.name}", heading_style),
        Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]),
        Spacer(1, 0.5 * cm),
        Paragraph(
            f"Location: {site.latitude:.4f}, {site.longitude:.4f} &nbsp;&nbsp; "
            f"Habitat: {site.habitat_type.value} &nbsp;&nbsp; "
            f"Protected Area: {site.protected_area or 'N/A'}",
            styles["Normal"],
        ),
        Spacer(1, 0.5 * cm),
    ]

    if data["biodiversity"]:
        b = data["biodiversity"]
        elements.append(Paragraph("Biodiversity Assessment", heading_style))
        bio_table = Table(
            [
                ["Metric", "Value"],
                ["Overall Ecosystem Health Score", f"{b.overall_ecosystem_health_score:.1f}/100"],
                ["Conservation Status", b.conservation_status_label],
                ["Species Richness", f"{b.species_richness:.0f}"],
                ["Shannon Diversity Index", f"{b.shannon_diversity_index:.3f}"],
            ],
            colWidths=[9 * cm, 6 * cm],
        )
        bio_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1C4D25")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        elements.append(bio_table)
        elements.append(Spacer(1, 0.5 * cm))

    if data["habitat"]:
        h = data["habitat"]
        elements.append(Paragraph("Habitat Assessment", heading_style))
        habitat_table = Table(
            [
                ["Metric", "Value"],
                ["Habitat Quality Score", f"{h.habitat_quality_score:.1f}/100"],
                ["Degradation Status", h.degradation_status_label],
                ["Vegetation Index (proxy)", f"{h.vegetation_index_proxy:.1f}/100"],
            ],
            colWidths=[9 * cm, 6 * cm],
        )
        habitat_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6B9080")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        elements.append(habitat_table)
        elements.append(Spacer(1, 0.5 * cm))

    if data["population_estimates"]:
        elements.append(Paragraph("Population Estimates", heading_style))
        rows = [["Species", "Est. Size", "Trend"]]
        seen = set()
        for p in data["population_estimates"]:
            if p.species_common_name in seen:
                continue
            seen.add(p.species_common_name)
            rows.append([p.species_common_name, f"{p.estimated_population_size:.0f}", p.trend_label])
        pop_table = Table(rows, colWidths=[7 * cm, 4 * cm, 4 * cm])
        pop_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1C4D25")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        elements.append(pop_table)
        elements.append(Spacer(1, 0.5 * cm))

    if data["recommendations"]:
        elements.append(Paragraph("Conservation Recommendations", heading_style))
        for rec in data["recommendations"][:10]:
            elements.append(Paragraph(f"<b>[{rec.priority.value.upper()}]</b> {rec.title}", styles["Normal"]))
            elements.append(Paragraph(rec.description, styles["Normal"]))
            elements.append(Spacer(1, 0.2 * cm))

    elements.append(Spacer(1, 0.5 * cm))
    elements.append(Paragraph(f"Total species observations on record: {len(data['observations'])}", styles["Normal"]))

    doc.build(elements)
    buffer.seek(0)

    filename = f"{site.name.replace(' ', '_')}_report.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{monitoring_site_id}/excel")
def export_site_report_excel(
    monitoring_site_id: str,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    data = _gather_site_report_data(db, monitoring_site_id)
    site = data["site"]

    wb = Workbook()
    header_fill = PatternFill(start_color="1C4D25", end_color="1C4D25", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # --- Sheet 1: Species Observations ---
    ws = wb.active
    ws.title = "Species Observations"
    headers = ["Species", "Scientific Name", "Group", "Conservation Status", "Confidence", "Source", "Detected At"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for obs in data["observations"]:
        ws.append([
            obs.species_common_name,
            obs.species_scientific_name or "",
            obs.species_group.value,
            obs.conservation_status.value,
            obs.confidence_score,
            obs.media_asset.source_type.value if obs.media_asset else "unknown",
            obs.detected_at.strftime("%Y-%m-%d %H:%M"),
        ])
    for col in ws.columns:
        max_len = max(len(str(c.value)) for c in col if c.value is not None) if any(c.value for c in col) else 10
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # --- Sheet 2: Population Estimates ---
    ws2 = wb.create_sheet("Population Estimates")
    ws2.append(["Species", "Estimated Size", "Density (per km²)", "Growth Rate %", "Trend", "Assessed At"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    for p in data["population_estimates"]:
        ws2.append([
            p.species_common_name, p.estimated_population_size, p.population_density,
            p.growth_rate_percent, p.trend_label, p.assessed_at.strftime("%Y-%m-%d %H:%M"),
        ])

    # --- Sheet 3: Conservation Recommendations ---
    ws3 = wb.create_sheet("Recommendations")
    ws3.append(["Priority", "Category", "Title", "Description", "Status", "Generated At"])
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
    for rec in data["recommendations"]:
        ws3.append([
            rec.priority.value, rec.category.value, rec.title, rec.description,
            rec.is_resolved, rec.generated_at.strftime("%Y-%m-%d %H:%M"),
        ])
    for col in ws3.columns:
        ws3.column_dimensions[col[0].column_letter].width = 30

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{site.name.replace(' ', '_')}_report.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

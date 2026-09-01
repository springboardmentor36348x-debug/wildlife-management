"""
Reports & Export System Engine (Milestone 4, Feature 5).
Generates PDF and Excel reports for all 5 report types using ReportLab and openpyxl,
pulling authentic data live from the database and intelligence engines.
"""
import io
import os
import uuid
from datetime import datetime, timezone
from typing import Any

from sqlalchemy import func
from sqlalchemy.orm import Session

# ReportLab for PDF generation
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether, HRFlowable
)
from reportlab.pdfgen import canvas

# openpyxl for Excel generation
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.models.user import User
from app.models.survey import Survey, MonitoringSite
from app.models.observation import Observation, ObservationType
from app.models.incident import (
    Incident, GeneratedReport, ReportFormat, ReportType,
    RestorationActionRecord, ActionStatus
)
from app.schemas.report import ReportFilterParams
from app.services import (
    population_service,
    habitat_service,
    conservation_service,
    health_score_service,
)

REPORT_TYPE_METADATA = {
    ReportType.WILDLIFE_SURVEY: {
        "name": "Wildlife Survey Report",
        "description": "Comprehensive survey tracking, monitoring site device status, and raw observation telemetry.",
        "supported_formats": ["pdf", "excel"],
        "suggested_filters": ["survey_id", "site_id", "start_date", "end_date"],
    },
    ReportType.SPECIES_POPULATION: {
        "name": "Species Population Report",
        "description": "Species detection counts, relative density proxy, longitudinal trend data, and geographic distribution.",
        "supported_formats": ["pdf", "excel"],
        "suggested_filters": ["survey_id", "species", "start_date", "end_date"],
    },
    ReportType.BIODIVERSITY: {
        "name": "Biodiversity Assessment Report",
        "description": "Ecosystem health scores, 5-component weighted breakdown, species richness, and conservation status.",
        "supported_formats": ["pdf", "excel"],
        "suggested_filters": ["survey_id", "site_id", "protected_area"],
    },
    ReportType.HABITAT_ASSESSMENT: {
        "name": "Habitat Assessment Report",
        "description": "Habitat type classification, vegetation & degradation indicators, and species suitability modeling.",
        "supported_formats": ["pdf", "excel"],
        "suggested_filters": ["site_id", "species"],
    },
    ReportType.CONSERVATION: {
        "name": "Conservation Action & Priorities Report",
        "description": "Ranked conservation priorities, active threat alerts, restoration recommendations, and field incidents.",
        "supported_formats": ["pdf", "excel"],
        "suggested_filters": ["survey_id", "site_id"],
    },
}


class NumberedCanvas(canvas.Canvas):
    """Adds 'Page X of Y' and header/footer rules to every page of the PDF."""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count: int):
        self.saveState()
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#4a5568"))

        # Header (pages > 1)
        if self._pageNumber > 1:
            self.drawString(54, 11 * inch - 36, "Wildlife Population Intelligence System — Official Report")
            self.setStrokeColor(colors.HexColor("#cbd5e1"))
            self.setLineWidth(0.5)
            self.line(54, 11 * inch - 42, 8.5 * inch - 54, 11 * inch - 42)

        # Footer
        self.setStrokeColor(colors.HexColor("#cbd5e1"))
        self.setLineWidth(0.5)
        self.line(54, 48, 8.5 * inch - 54, 48)
        self.drawString(54, 34, "CONFIDENTIAL & PROPRIETARY — FOR CONSERVATION MANAGEMENT USE ONLY")
        page_str = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(8.5 * inch - 54, 34, page_str)
        self.restoreState()


def _gather_report_data(db: Session, report_type: ReportType, filters: ReportFilterParams | None) -> dict[str, Any]:
    """Pulls and structures real live data from the database and analysis engines."""
    filters = filters or ReportFilterParams()
    data: dict[str, Any] = {
        "report_type": report_type,
        "generated_at": datetime.now(timezone.utc),
        "filters": filters.model_dump(exclude_none=True),
    }

    # Base query helpers
    obs_query = db.query(Observation)
    site_query = db.query(MonitoringSite)
    survey_query = db.query(Survey)

    if filters.survey_id:
        survey_query = survey_query.filter(Survey.id == filters.survey_id)
        site_query = site_query.filter(MonitoringSite.survey_id == filters.survey_id)
        obs_query = obs_query.join(MonitoringSite, Observation.site_id == MonitoringSite.id).filter(
            MonitoringSite.survey_id == filters.survey_id
        )
    if filters.site_id:
        site_query = site_query.filter(MonitoringSite.id == filters.site_id)
        obs_query = obs_query.filter(Observation.site_id == filters.site_id)
    if filters.species:
        obs_query = obs_query.filter(Observation.species_label == filters.species)
    if filters.start_date:
        obs_query = obs_query.filter(Observation.captured_at >= filters.start_date)
    if filters.end_date:
        obs_query = obs_query.filter(Observation.captured_at <= filters.end_date)
    if filters.protected_area:
        site_query = site_query.filter(MonitoringSite.protected_area == filters.protected_area)

    sites = site_query.all()
    surveys = survey_query.all()
    observations = obs_query.all()

    data["sites_count"] = len(sites)
    data["surveys_count"] = len(surveys)
    data["observations_count"] = len(observations)

    if report_type == ReportType.WILDLIFE_SURVEY:
        surveys_list = []
        for s in surveys:
            site_count = len(s.monitoring_sites)
            obs_count = sum(len(site.observations) for site in s.monitoring_sites)
            surveys_list.append({
                "id": s.id[:8],
                "name": s.name,
                "status": s.status.value if hasattr(s.status, "value") else str(s.status),
                "protected_area": s.protected_area or "Unassigned",
                "start_date": s.start_date.strftime("%Y-%m-%d") if s.start_date else "—",
                "sites_count": site_count,
                "observations_count": obs_count,
            })
        sites_list = []
        for site in sites:
            sites_list.append({
                "id": site.id[:8],
                "name": site.site_name,
                "survey": site.survey.name if site.survey else "—",
                "habitat": site.habitat_type.value if hasattr(site.habitat_type, "value") else str(site.habitat_type),
                "device": site.monitoring_device.value if hasattr(site.monitoring_device, "value") else str(site.monitoring_device),
                "latitude": round(site.latitude, 4),
                "longitude": round(site.longitude, 4),
                "observations_count": len(site.observations),
                "is_active": "Active" if site.is_active == "true" else "Inactive",
            })
        data["summary_metrics"] = {
            "Total Surveys": len(surveys),
            "Active Surveys": sum(1 for s in surveys if (s.status.value if hasattr(s.status, "value") else str(s.status)) == "active"),
            "Total Monitoring Sites": len(sites),
            "Total Field Observations": len(observations),
        }
        data["surveys_table"] = surveys_list
        data["sites_table"] = sites_list

    elif report_type == ReportType.SPECIES_POPULATION:
        species_counts = population_service.get_population_counts(db, survey_id=filters.survey_id)
        density_data = population_service.get_population_density(db, survey_id=filters.survey_id)
        distribution_data = population_service.get_species_distribution(db, survey_id=filters.survey_id)

        top_species = species_counts[0]["species"] if species_counts else "None"
        total_detections = sum(r["count"] for r in species_counts)

        data["summary_metrics"] = {
            "Total Confirmed Species": len(species_counts),
            "Total Detections Logged": total_detections,
            "Predominant Species": top_species.capitalize(),
            "Monitored Sites with Detections": len({d["site_id"] for d in density_data}),
        }
        data["species_counts"] = species_counts
        data["density_table"] = density_data
        data["distribution_table"] = distribution_data

    elif report_type == ReportType.BIODIVERSITY:
        health_scores = health_score_service.calculate_ecosystem_health_all_sites(db)
        if filters.site_id:
            health_scores = [h for h in health_scores if h.get("site_id") == filters.site_id]

        status_counts = {"Excellent": 0, "Healthy": 0, "Moderate Concern": 0, "Vulnerable": 0, "Critical": 0}
        avg_score = 0
        if health_scores:
            avg_score = round(sum(h["ecosystem_health_score"] for h in health_scores) / len(health_scores), 1)
            for h in health_scores:
                status = h.get("conservation_status", "Critical")
                if status in status_counts:
                    status_counts[status] += 1

        sites_by_id = {s.id: s for s in db.query(MonitoringSite).all()}
        health_table = []
        for h in health_scores:
            site = sites_by_id.get(h.get("site_id"))
            components = h.get("components", {})
            health_table.append({
                "site_name": site.site_name if site else (h.get("site_id") or "Unknown")[:8],
                "score": h.get("ecosystem_health_score", 0),
                "status": h.get("conservation_status", "Unknown"),
                "diversity_score": components.get("species_diversity", {}).get("score", 0),
                "stability_score": components.get("population_stability", {}).get("score", 0),
                "habitat_score": components.get("habitat_quality", {}).get("score", 0),
                "endangered_score": components.get("endangered_species_status", {}).get("score", 0),
            })

        data["summary_metrics"] = {
            "Average Health Score": f"{avg_score}/100",
            "Monitored Sites Scored": len(health_scores),
            "Healthy / Excellent Sites": status_counts["Healthy"] + status_counts["Excellent"],
            "Threatened (Vuln/Crit)": status_counts["Vulnerable"] + status_counts["Critical"],
        }
        data["status_breakdown"] = status_counts
        data["health_table"] = health_table

    elif report_type == ReportType.HABITAT_ASSESSMENT:
        habitat_table = []
        habitat_counts: dict[str, int] = {}
        degradation_counts = {"stable": 0, "declining": 0, "insufficient_data": 0}

        for site in sites:
            htype = habitat_service.classify_habitat(site)
            habitat_counts[htype] = habitat_counts.get(htype, 0) + 1
            deg = habitat_service.detect_habitat_degradation(db, site_id=site.id)
            status = deg.get("status", "insufficient_data")
            if status in degradation_counts:
                degradation_counts[status] += 1
            
            # Predict suitability for top species or elephant/bird
            test_species = filters.species or "elephant"
            suit = habitat_service.predict_habitat_suitability(db, site_id=site.id, species_label=test_species)

            habitat_table.append({
                "site_name": site.site_name,
                "habitat_type": htype.capitalize(),
                "degradation_status": status.replace("_", " ").capitalize(),
                "change_pct": f"{deg.get('change_pct')}%" if deg.get("change_pct") is not None else "N/A",
                "recent_obs": deg.get("recent_count", 0),
                "suitability_species": test_species.capitalize(),
                "suitability_score": suit.get("suitability_score", 0),
            })

        data["summary_metrics"] = {
            "Total Sites Monitored": len(sites),
            "Primary Habitat Types": len(habitat_counts),
            "Stable Habitat Sites": degradation_counts["stable"],
            "Declining Activity Flagged": degradation_counts["declining"],
        }
        data["habitat_counts"] = habitat_counts
        data["habitat_table"] = habitat_table

    elif report_type == ReportType.CONSERVATION:
        priorities = conservation_service.get_conservation_priorities(db)
        monitoring_opt = conservation_service.optimize_monitoring(db)
        allocations = conservation_service.recommend_resource_allocation(db)
        incidents = db.query(Incident).order_by(Incident.reported_at.desc()).all()

        high_priority = sum(1 for p in priorities if p.get("priority") == "high")
        med_priority = sum(1 for p in priorities if p.get("priority") == "medium")
        open_incidents = sum(1 for i in incidents if i.status == "open")

        data["summary_metrics"] = {
            "High Priority Sites": high_priority,
            "Medium Priority Sites": med_priority,
            "Total Monitored Sites": len(priorities),
            "Active Field Incidents": open_incidents,
        }
        data["priorities_table"] = priorities
        data["allocations_table"] = allocations
        data["incidents_table"] = [
            {
                "id": inc.id[:8],
                "title": inc.title,
                "type": inc.incident_type.value if hasattr(inc.incident_type, "value") else str(inc.incident_type),
                "severity": inc.severity.value if hasattr(inc.severity, "value") else str(inc.severity),
                "status": inc.status.value if hasattr(inc.status, "value") else str(inc.status),
                "reported_at": inc.reported_at.strftime("%Y-%m-%d %H:%M") if inc.reported_at else "—",
            }
            for inc in incidents[:15]
        ]

    return data


def _generate_pdf(report_title: str, report_type: ReportType, data: dict[str, Any], generated_by_user: User) -> bytes:
    """Renders a PDF report via ReportLab."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=54,
        rightMargin=54,
        topMargin=54,
        bottomMargin=54,
    )

    styles = getSampleStyleSheet()

    # Custom styles
    primary_color = colors.HexColor("#1b4332")
    secondary_color = colors.HexColor("#2d6a4f")
    text_dark = colors.HexColor("#1a202c")
    text_muted = colors.HexColor("#4a5568")
    accent_gold = colors.HexColor("#d97706")

    title_style = ParagraphStyle(
        "DocTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=20,
        leading=24,
        textColor=primary_color,
        spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        "DocSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        leading=14,
        textColor=text_muted,
        spaceAfter=14,
    )
    section_heading = ParagraphStyle(
        "SectionHeading",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=13,
        leading=17,
        textColor=secondary_color,
        spaceBefore=12,
        spaceAfter=8,
        keepWithNext=True,
    )
    body_style = ParagraphStyle(
        "Body",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=text_dark,
    )
    meta_label = ParagraphStyle(
        "MetaLabel",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=text_muted,
    )
    meta_val = ParagraphStyle(
        "MetaVal",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        textColor=text_dark,
    )

    story = []

    # Title & Metadata Header Box
    story.append(Paragraph(report_title, title_style))
    story.append(Paragraph(REPORT_TYPE_METADATA.get(report_type, {}).get("description", "Conservation Report"), subtitle_style))

    # Meta banner table
    meta_data = [
        [
            Paragraph("GENERATED DATE", meta_label),
            Paragraph("GENERATED BY", meta_label),
            Paragraph("REPORT TYPE", meta_label),
            Paragraph("SECURITY CLASSIFICATION", meta_label),
        ],
        [
            Paragraph(datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"), meta_val),
            Paragraph(f"{generated_by_user.full_name} ({generated_by_user.role.value})", meta_val),
            Paragraph(REPORT_TYPE_METADATA.get(report_type, {}).get("name", "Report"), meta_val),
            Paragraph("Official Conservation Record", meta_val),
        ],
    ]
    meta_table = Table(meta_data, colWidths=[120, 150, 140, 110])
    meta_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 14))

    # Section 1: Executive KPI Summary
    story.append(Paragraph("Executive Summary & Key Metrics", section_heading))
    metrics = data.get("summary_metrics", {})
    if metrics:
        metric_items = list(metrics.items())
        # Render as a 2 or 4 column grid table
        grid_rows = []
        for i in range(0, len(metric_items), 2):
            pair = metric_items[i:i+2]
            row_labels = []
            row_vals = []
            for label, val in pair:
                row_labels.append(Paragraph(str(label).upper(), ParagraphStyle("KpiL", fontName="Helvetica-Bold", fontSize=8, textColor=colors.HexColor("#4b5563"))))
                row_vals.append(Paragraph(str(val), ParagraphStyle("KpiV", fontName="Helvetica-Bold", fontSize=14, textColor=primary_color)))
            if len(pair) == 1:
                row_labels.extend(["", ""])
                row_vals.extend(["", ""])
            grid_rows.append(row_labels)
            grid_rows.append(row_vals)

        kpi_table = Table(grid_rows, colWidths=[250, 250])
        kpi_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f0fdf4")),
            ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#bbf7d0")),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dcfce7")),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 12),
            ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ]))
        story.append(kpi_table)
    story.append(Spacer(1, 14))

    # Section 2: Detailed Section Tables based on report type
    th_style = ParagraphStyle("TH", fontName="Helvetica-Bold", fontSize=8, textColor=colors.white)
    td_style = ParagraphStyle("TD", fontName="Helvetica", fontSize=8, textColor=text_dark)

    if report_type == ReportType.WILDLIFE_SURVEY:
        story.append(Paragraph("Monitoring Surveys Registry", section_heading))
        surveys_table = data.get("surveys_table", [])
        if surveys_table:
            table_data = [[
                Paragraph("SURVEY ID", th_style),
                Paragraph("NAME", th_style),
                Paragraph("STATUS", th_style),
                Paragraph("PROTECTED AREA", th_style),
                Paragraph("SITES", th_style),
                Paragraph("OBSERVATIONS", th_style),
            ]]
            for s in surveys_table[:12]:
                table_data.append([
                    Paragraph(s["id"], td_style),
                    Paragraph(s["name"], td_style),
                    Paragraph(s["status"].capitalize(), td_style),
                    Paragraph(s["protected_area"], td_style),
                    Paragraph(str(s["sites_count"]), td_style),
                    Paragraph(str(s["observations_count"]), td_style),
                ])
            t = Table(table_data, colWidths=[65, 140, 65, 110, 50, 70])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), secondary_color),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(t)

        story.append(Spacer(1, 14))
        story.append(Paragraph("Monitoring Sites Infrastructure", section_heading))
        sites_table = data.get("sites_table", [])
        if sites_table:
            table_data = [[
                Paragraph("SITE NAME", th_style),
                Paragraph("HABITAT", th_style),
                Paragraph("DEVICE", th_style),
                Paragraph("GPS LAT/LON", th_style),
                Paragraph("OBS COUNT", th_style),
                Paragraph("STATE", th_style),
            ]]
            for st in sites_table[:15]:
                table_data.append([
                    Paragraph(st["name"], td_style),
                    Paragraph(st["habitat"].capitalize(), td_style),
                    Paragraph(st["device"].replace("_", " ").capitalize(), td_style),
                    Paragraph(f"{st['latitude']}, {st['longitude']}", td_style),
                    Paragraph(str(st["observations_count"]), td_style),
                    Paragraph(st["is_active"], td_style),
                ])
            t = Table(table_data, colWidths=[120, 80, 95, 105, 55, 45])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), secondary_color),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(t)

    elif report_type == ReportType.SPECIES_POPULATION:
        story.append(Paragraph("Species Occurrence & Detection Counts", section_heading))
        species_counts = data.get("species_counts", [])
        if species_counts:
            table_data = [[
                Paragraph("SPECIES", th_style),
                Paragraph("TOTAL DETECTIONS", th_style),
                Paragraph("RELATIVE SHARE", th_style),
            ]]
            total_d = sum(r["count"] for r in species_counts) or 1
            for r in species_counts[:15]:
                pct = round((r["count"] / total_d) * 100, 1)
                table_data.append([
                    Paragraph(r["species"].capitalize(), td_style),
                    Paragraph(str(r["count"]), td_style),
                    Paragraph(f"{pct}%", td_style),
                ])
            t = Table(table_data, colWidths=[200, 150, 150])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), secondary_color),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(t)

        story.append(Spacer(1, 14))
        story.append(Paragraph("Site Population Density Proxy", section_heading))
        density_table = data.get("density_table", [])
        if density_table:
            table_data = [[
                Paragraph("SITE NAME", th_style),
                Paragraph("SPECIES", th_style),
                Paragraph("OBSERVATION COUNT", th_style),
            ]]
            for row in density_table[:15]:
                table_data.append([
                    Paragraph(row["site_name"], td_style),
                    Paragraph(row["species"].capitalize(), td_style),
                    Paragraph(str(row["count"]), td_style),
                ])
            t = Table(table_data, colWidths=[200, 150, 150])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), secondary_color),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(t)

    elif report_type == ReportType.BIODIVERSITY:
        story.append(Paragraph("Ecosystem Health & Diversity Scores per Site", section_heading))
        health_table = data.get("health_table", [])
        if health_table:
            table_data = [[
                Paragraph("SITE NAME", th_style),
                Paragraph("HEALTH SCORE", th_style),
                Paragraph("STATUS", th_style),
                Paragraph("DIVERSITY", th_style),
                Paragraph("STABILITY", th_style),
                Paragraph("HABITAT", th_style),
                Paragraph("ENDANGERED", th_style),
            ]]
            for row in health_table[:18]:
                table_data.append([
                    Paragraph(row["site_name"], td_style),
                    Paragraph(f"{row['score']}/100", ParagraphStyle("Sc", fontName="Helvetica-Bold", fontSize=8, textColor=primary_color)),
                    Paragraph(row["status"], td_style),
                    Paragraph(str(row["diversity_score"]), td_style),
                    Paragraph(str(row["stability_score"]), td_style),
                    Paragraph(str(row["habitat_score"]), td_style),
                    Paragraph(str(row["endangered_score"]), td_style),
                ])
            t = Table(table_data, colWidths=[120, 65, 80, 55, 55, 55, 70])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), secondary_color),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(t)

    elif report_type == ReportType.HABITAT_ASSESSMENT:
        story.append(Paragraph("Habitat Classification & Degradation Assessment", section_heading))
        habitat_table = data.get("habitat_table", [])
        if habitat_table:
            table_data = [[
                Paragraph("SITE NAME", th_style),
                Paragraph("HABITAT TYPE", th_style),
                Paragraph("DEGRADATION", th_style),
                Paragraph("DELTA %", th_style),
                Paragraph("RECENT OBS", th_style),
                Paragraph("SUITABILITY", th_style),
            ]]
            for row in habitat_table[:18]:
                table_data.append([
                    Paragraph(row["site_name"], td_style),
                    Paragraph(row["habitat_type"], td_style),
                    Paragraph(row["degradation_status"], td_style),
                    Paragraph(row["change_pct"], td_style),
                    Paragraph(str(row["recent_obs"]), td_style),
                    Paragraph(f"{row['suitability_species']}: {row['suitability_score']}/100", td_style),
                ])
            t = Table(table_data, colWidths=[110, 80, 85, 55, 65, 105])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), secondary_color),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(t)

    elif report_type == ReportType.CONSERVATION:
        story.append(Paragraph("Ranked Conservation Priorities", section_heading))
        priorities = data.get("priorities_table", [])
        if priorities:
            table_data = [[
                Paragraph("SITE NAME", th_style),
                Paragraph("PRIORITY LEVEL", th_style),
                Paragraph("ASSESSMENT & JUSTIFICATION", th_style),
            ]]
            for p in priorities[:12]:
                table_data.append([
                    Paragraph(p["site_name"], td_style),
                    Paragraph(p["priority"].upper(), ParagraphStyle("Pr", fontName="Helvetica-Bold", fontSize=8, textColor=colors.HexColor("#b91c1c") if p["priority"] == "high" else colors.HexColor("#b45309") if p["priority"] == "medium" else colors.HexColor("#15803d"))),
                    Paragraph(p["reasoning"], td_style),
                ])
            t = Table(table_data, colWidths=[120, 80, 300])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), secondary_color),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(t)

        story.append(Spacer(1, 14))
        story.append(Paragraph("Recent Field Incidents", section_heading))
        incidents = data.get("incidents_table", [])
        if incidents:
            table_data = [[
                Paragraph("INCIDENT ID", th_style),
                Paragraph("TITLE", th_style),
                Paragraph("TYPE", th_style),
                Paragraph("SEVERITY", th_style),
                Paragraph("STATUS", th_style),
                Paragraph("TIMESTAMP", th_style),
            ]]
            for inc in incidents[:10]:
                table_data.append([
                    Paragraph(inc["id"], td_style),
                    Paragraph(inc["title"], td_style),
                    Paragraph(inc["type"].replace("_", " ").capitalize(), td_style),
                    Paragraph(inc["severity"].capitalize(), td_style),
                    Paragraph(inc["status"].capitalize(), td_style),
                    Paragraph(inc["reported_at"], td_style),
                ])
            t = Table(table_data, colWidths=[65, 130, 85, 60, 60, 100])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), secondary_color),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(t)

    # Build the document
    doc.build(story, canvasmaker=NumberedCanvas)
    buffer.seek(0)
    return buffer.getvalue()


def _generate_excel(report_title: str, report_type: ReportType, data: dict[str, Any], generated_by_user: User) -> bytes:
    """Renders a multi-sheet Excel report via openpyxl."""
    wb = openpyxl.Workbook()

    # Styling helpers
    header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True, color="1B4332")
    meta_font = Font(name="Calibri", size=10, italic=True, color="4A5568")
    bold_font = Font(name="Calibri", size=11, bold=True)
    kpi_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # ---- Sheet 1: Executive Summary ----
    ws_sum = wb.active
    ws_sum.title = "Executive Summary"
    ws_sum.views.sheetView[0].showGridLines = True

    ws_sum.append([report_title])
    ws_sum.cell(row=1, column=1).font = title_font
    ws_sum.append([REPORT_TYPE_METADATA.get(report_type, {}).get("description", "")])
    ws_sum.cell(row=2, column=1).font = meta_font
    ws_sum.append([])

    ws_sum.append(["Report Metadata", ""])
    ws_sum.cell(row=4, column=1).font = bold_font
    ws_sum.append(["Generated Date:", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")])
    ws_sum.append(["Generated By:", f"{generated_by_user.full_name} ({generated_by_user.role.value})"])
    ws_sum.append(["Report Type:", REPORT_TYPE_METADATA.get(report_type, {}).get("name", "")])
    ws_sum.append(["Security Classification:", "Official Conservation Record"])
    ws_sum.append([])

    ws_sum.append(["Key Performance Indicators", "Value"])
    ws_sum.cell(row=10, column=1).font = header_font
    ws_sum.cell(row=10, column=1).fill = header_fill
    ws_sum.cell(row=10, column=2).font = header_font
    ws_sum.cell(row=10, column=2).fill = header_fill

    row_idx = 11
    for k, v in data.get("summary_metrics", {}).items():
        ws_sum.append([k, str(v)])
        ws_sum.cell(row=row_idx, column=1).border = thin_border
        ws_sum.cell(row=row_idx, column=2).border = thin_border
        ws_sum.cell(row=row_idx, column=1).fill = kpi_fill
        ws_sum.cell(row=row_idx, column=2).fill = kpi_fill
        ws_sum.cell(row=row_idx, column=2).font = bold_font
        row_idx += 1

    # ---- Sheet 2: Detailed Data ----
    ws_det = wb.create_sheet(title="Detailed Records")
    ws_det.views.sheetView[0].showGridLines = True

    if report_type == ReportType.WILDLIFE_SURVEY:
        headers = ["Survey ID", "Survey Name", "Status", "Protected Area", "Sites Count", "Observation Count"]
        ws_det.append(headers)
        for s in data.get("surveys_table", []):
            ws_det.append([s["id"], s["name"], s["status"], s["protected_area"], s["sites_count"], s["observations_count"]])

    elif report_type == ReportType.SPECIES_POPULATION:
        headers = ["Species", "Observation Count", "Relative Share %"]
        ws_det.append(headers)
        tot = sum(r["count"] for r in data.get("species_counts", [])) or 1
        for r in data.get("species_counts", []):
            ws_det.append([r["species"], r["count"], round((r["count"] / tot) * 100, 2)])

    elif report_type == ReportType.BIODIVERSITY:
        headers = ["Site Name", "Health Score (0-100)", "Status", "Diversity Score", "Stability Score", "Habitat Score", "Endangered Status Score"]
        ws_det.append(headers)
        for h in data.get("health_table", []):
            ws_det.append([h["site_name"], h["score"], h["status"], h["diversity_score"], h["stability_score"], h["habitat_score"], h["endangered_score"]])

    elif report_type == ReportType.HABITAT_ASSESSMENT:
        headers = ["Site Name", "Habitat Type", "Degradation Status", "Trend Delta %", "Recent Observations", "Species Suitability Score"]
        ws_det.append(headers)
        for ht in data.get("habitat_table", []):
            ws_det.append([ht["site_name"], ht["habitat_type"], ht["degradation_status"], ht["change_pct"], ht["recent_obs"], ht["suitability_score"]])

    elif report_type == ReportType.CONSERVATION:
        headers = ["Site Name", "Priority Level", "Assessment & Reasoning"]
        ws_det.append(headers)
        for p in data.get("priorities_table", []):
            ws_det.append([p["site_name"], p["priority"], p["reasoning"]])

    # Style Header Row on Detailed Data
    for col in range(1, ws_det.max_column + 1):
        cell = ws_det.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in range(2, ws_det.max_row + 1):
        for c in range(1, ws_det.max_column + 1):
            ws_det.cell(row=r, column=c).border = thin_border

    # ---- Sheet 3: Site & Geography Metadata ----
    ws_geo = wb.create_sheet(title="Sites & Geospatial Data")
    ws_geo.views.sheetView[0].showGridLines = True
    geo_headers = ["Site ID", "Site Name", "Survey ID", "Latitude", "Longitude", "Habitat Type", "Monitoring Device", "Protected Area", "Active State"]
    ws_geo.append(geo_headers)
    for col in range(1, len(geo_headers) + 1):
        c = ws_geo.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill

    sites = data.get("sites_table", [])
    if not sites:
        # fallback to direct query if not in sites_table
        sites = [
            {
                "id": s["site_id"] if "site_id" in s else s.get("id", ""),
                "name": s["site_name"] if "site_name" in s else s.get("name", ""),
                "survey": s.get("survey", ""),
                "latitude": s.get("latitude", 0),
                "longitude": s.get("longitude", 0),
                "habitat": s.get("habitat", ""),
                "device": s.get("device", ""),
                "protected_area": s.get("protected_area", ""),
                "is_active": s.get("is_active", "Active"),
            }
            for s in data.get("sites_table", [])
        ]
    for s in sites:
        ws_geo.append([
            s.get("id", ""),
            s.get("name", ""),
            s.get("survey", ""),
            s.get("latitude", 0),
            s.get("longitude", 0),
            s.get("habitat", ""),
            s.get("device", ""),
            s.get("protected_area", ""),
            s.get("is_active", "Active"),
        ])

    for r in range(2, ws_geo.max_row + 1):
        for c in range(1, ws_geo.max_column + 1):
            ws_geo.cell(row=r, column=c).border = thin_border

    # Auto-adjust column widths across all sheets
    for sheet in [ws_sum, ws_det, ws_geo]:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_report(
    db: Session,
    report_type: ReportType,
    file_format: ReportFormat,
    generated_by_user: User,
    filters: ReportFilterParams | None = None,
    title: str | None = None,
) -> GeneratedReport:
    """Generates a complete report, saves the binary file to disk, and records it in the database."""
    report_id = str(uuid.uuid4())
    default_title = f"{REPORT_TYPE_METADATA.get(report_type, {}).get('name', 'Conservation Report')} — {datetime.now(timezone.utc).strftime('%b %d, %Y')}"
    final_title = title or default_title

    # Pull live data
    data = _gather_report_data(db, report_type=report_type, filters=filters)

    # Generate binary content
    if file_format == ReportFormat.PDF:
        file_bytes = _generate_pdf(final_title, report_type, data, generated_by_user)
        ext = ".pdf"
    else:
        file_bytes = _generate_excel(final_title, report_type, data, generated_by_user)
        ext = ".xlsx"

    # Store on disk under uploads/reports
    reports_dir = os.path.join(settings.UPLOAD_DIR, "reports")
    os.makedirs(reports_dir, exist_ok=True)
    filename = f"{report_id}{ext}"
    disk_path = os.path.join(reports_dir, filename)

    with open(disk_path, "wb") as f:
        f.write(file_bytes)

    download_url = f"{settings.API_V1_PREFIX}/reports/{report_id}/download"

    report_record = GeneratedReport(
        id=report_id,
        title=final_title,
        report_type=report_type,
        file_format=file_format,
        file_path=disk_path,
        file_size_bytes=len(file_bytes),
        download_url=download_url,
        filters_json=filters.model_dump(exclude_none=True) if filters else {},
        summary_metrics=data.get("summary_metrics", {}),
        generated_by=generated_by_user.id,
        created_at=datetime.now(timezone.utc),
        download_count=0,
    )
    db.add(report_record)
    db.commit()
    db.refresh(report_record)
    return report_record


def get_report_types_metadata() -> list[dict[str, Any]]:
    """Returns metadata for all supported report types and formats."""
    result = []
    for rtype, meta in REPORT_TYPE_METADATA.items():
        result.append({
            "type": rtype.value,
            "name": meta["name"],
            "description": meta["description"],
            "supported_formats": meta["supported_formats"],
            "suggested_filters": meta["suggested_filters"],
        })
    return result

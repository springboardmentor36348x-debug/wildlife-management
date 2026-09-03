"""
Wildlife Intelligence Excel Report Generator
Exports observation logs, population records, biodiversity metrics, and survey data to Excel spreadsheets.
"""

import os
from datetime import datetime
from typing import Dict, Any, List
import logging

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("OpenPyXL not installed; fallback CSV/Text export active.")


class WildlifeExcelGenerator:
    """Generates styled Excel reports for wildlife observations and ecological analytics"""

    @staticmethod
    def generate_excel_report(
        report_title: str,
        site_data: Dict[str, Any],
        observations_data: List[Dict[str, Any]],
        health_data: Dict[str, Any],
        recommendations: List[Dict[str, Any]],
        output_filepath: str
    ) -> bool:
        """Create a multi-tab formatted Excel report"""
        os.makedirs(os.path.dirname(output_filepath), exist_ok=True)

        if not OPENPYXL_AVAILABLE:
            # Fallback CSV
            with open(output_filepath, "w", encoding="utf-8") as f:
                f.write("Species,Group,Observations,Count,IUCN_Status\n")
                for obs in observations_data:
                    f.write(f"{obs.get('species_name')},{obs.get('species_group')},{obs.get('obs_count')},{obs.get('individual_count')},{obs.get('is_endangered')}\n")
            return True

        try:
            wb = openpyxl.Workbook()

            # Styles
            header_fill = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid")
            sub_fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            title_font = Font(name="Calibri", size=16, bold=True, color="065F46")
            regular_font = Font(name="Calibri", size=10)
            bold_font = Font(name="Calibri", size=10, bold=True)
            border_side = Side(style='thin', color='D1D5DB')
            thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

            # TAB 1: Summary & Ecosystem Health
            ws_summary = wb.active
            ws_summary.title = "Ecosystem Health"
            ws_summary.views.sheetView[0].showGridLines = True

            ws_summary["A1"] = report_title
            ws_summary["A1"].font = title_font
            ws_summary["A2"] = f"Monitoring Site: {site_data.get('site_name', 'All Sites')} | Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
            ws_summary["A2"].font = Font(name="Calibri", size=10, italic=True, color="4B5563")

            headers_h = ["Dimension", "Score (0-100)", "Model Weight", "Status"]
            for col_idx, h in enumerate(headers_h, start=1):
                cell = ws_summary.cell(row=4, column=col_idx, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            health_rows = [
                ("Species Diversity", health_data.get("species_diversity_score", 85.0), "30%", "Optimal"),
                ("Population Stability", health_data.get("population_stability_score", 78.0), "25%", "Stable"),
                ("Habitat Quality", health_data.get("habitat_quality_score", 82.0), "20%", "Good"),
                ("Endangered Species Status", health_data.get("endangered_species_score", 90.0), "15%", "Monitored"),
                ("Environmental Conditions", health_data.get("environmental_conditions_score", 75.0), "10%", "Favorable"),
                ("Composite Health Score", health_data.get("overall_health_score", 82.5), "100%", health_data.get("health_status", "Healthy"))
            ]

            for row_idx, r in enumerate(health_rows, start=5):
                for col_idx, val in enumerate(r, start=1):
                    cell = ws_summary.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = bold_font if row_idx == 10 else regular_font
                    cell.border = thin_border
                    if col_idx in [2, 3]:
                        cell.alignment = Alignment(horizontal="center")

            # TAB 2: Observations Log
            ws_obs = wb.create_sheet(title="Observations")
            ws_obs.views.sheetView[0].showGridLines = True
            headers_obs = ["Species Common Name", "Species Group", "Observation Count", "Total Individuals", "IUCN Status", "Endangered Flag"]
            for col_idx, h in enumerate(headers_obs, start=1):
                cell = ws_obs.cell(row=1, column=col_idx, value=h)
                cell.fill = sub_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for row_idx, obs in enumerate(observations_data, start=2):
                row_data = [
                    obs.get("species_name", "Bengal Tiger"),
                    obs.get("species_group", "Mammal"),
                    obs.get("obs_count", 1),
                    obs.get("individual_count", 1),
                    "Endangered" if obs.get("is_endangered") else "Least Concern",
                    "YES" if obs.get("is_endangered") else "NO"
                ]
                for col_idx, val in enumerate(row_data, start=1):
                    cell = ws_obs.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = regular_font
                    cell.border = thin_border
                    if col_idx in [3, 4, 6]:
                        cell.alignment = Alignment(horizontal="center")

            # TAB 3: Recommendations
            ws_rec = wb.create_sheet(title="Conservation Strategies")
            ws_rec.views.sheetView[0].showGridLines = True
            headers_rec = ["Priority", "Strategy Title", "Evidence Base", "Action Plan"]
            for col_idx, h in enumerate(headers_rec, start=1):
                cell = ws_rec.cell(row=1, column=col_idx, value=h)
                cell.fill = header_fill
                cell.font = header_font

            for row_idx, rec in enumerate(recommendations, start=2):
                row_data = [
                    rec.get("priority", "HIGH"),
                    rec.get("title", ""),
                    rec.get("evidence", ""),
                    rec.get("description", "")
                ]
                for col_idx, val in enumerate(row_data, start=1):
                    cell = ws_rec.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = regular_font
                    cell.border = thin_border

            # Auto-adjust column widths
            for ws in [ws_summary, ws_obs, ws_rec]:
                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = openpyxl.utils.get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

            wb.save(output_filepath)
            return True
        except Exception as e:
            logger.error(f"Failed to generate Excel report: {e}")
            return False

excel_generator = WildlifeExcelGenerator()
